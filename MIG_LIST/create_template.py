"""Generate MIG_TAB_LIST Excel template with sample data."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MIG_TAB_LIST"

# --- Header style ---
header_font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
cell_font = Font(name="Consolas", size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

headers = ["OWNER", "TABLE_NAME", "WHERE_COL1", "PRE1", "WHERE_COL2", "PRE2"]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# --- Sample data ---
samples = [
    ("DBADM", "TEST_ORDER", "ORDER_DATE", "20250401", None, None),
    ("DBADM", "TEST_LOG", "LOG_DATE", "20250201", "LOG_DATE", "20250601"),
    ("DBADM", "TEST_CODE", None, None, None, None),
]

for row_idx, row_data in enumerate(samples, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border

# --- Column widths ---
widths = [15, 20, 15, 15, 15, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb.save("template/MIG_TAB_LIST_TEMPLATE.xlsx")
print("Template created: template/MIG_TAB_LIST_TEMPLATE.xlsx")

# ============================================================
# Sample Excel with realistic data
# ============================================================
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "MIG_TAB_LIST"

for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

sample_data = [
    ("SALES",  "ORDER_HEADER",  "ORDER_DATE",  "20250101", "ORDER_DATE",  "20250701"),
    ("SALES",  "ORDER_DETAIL",  "ORDER_DATE",  "20250101", "ORDER_DATE",  "20250701"),
    ("SALES",  "ORDER_STATUS",  None,           None,       None,          None),
    ("HR",     "EMPLOYEE",      None,           None,       None,          None),
    ("HR",     "SALARY_HIST",   "PAY_DATE",    "20250101", None,          None),
    ("HR",     "DEPT_CODE",     None,           None,       None,          None),
    ("INV",    "STOCK_MASTER",  None,           None,       None,          None),
    ("INV",    "STOCK_TRANS",   "TRANS_DATE",  "20250301", "TRANS_DATE",  "20250601"),
    ("INV",    "WAREHOUSE",     None,           None,       None,          None),
    ("FIN",    "GL_JOURNAL",    "POST_DATE",   "20250101", "POST_DATE",   "20260101"),
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border

for i, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb2.save("template/MIG_TAB_LIST_SAMPLE.xlsx")
print("Sample  created: template/MIG_TAB_LIST_SAMPLE.xlsx")
